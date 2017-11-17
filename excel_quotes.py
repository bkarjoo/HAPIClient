from openpyxl import *

excel_rows = dict()
quote_dictionary = dict()


def build_dictionary():
    try:
        wb = load_workbook('Book1.xlsm',True,True,True,False,True)
        ws = wb['Sheet1']

        first_row = 4
        i = 4
        symb = ws['A{0}'.format(i)].value
        while symb != None:
            excel_rows[symb] = i
            i += 1
            symb = ws['A{0}'.format(i)].value
        wb.close()
    except Exception as e:
        print e


def build_quote_dictionary():
    try:
        wb = load_workbook('Book1.xlsm',True,True,True,False,True)
        ws = wb['Sheet1']

        first_row = 4
        i = 4
        l = list()
        l.append(ws['A{0}'.format(i)].value)
        while l[0] != None:
            try:
                l.append(float(ws['E{0}'.format(i)].value))
                l.append(float(ws['F{0}'.format(i)].value))
                quote_dictionary[l[0]] = l
                i += 1
                l = list()
                l.append(ws['A{0}'.format(i)].value)
            except:
                i += 1
                l = list()
                l.append(ws['A{0}'.format(i)].value)


    except Exception as e:
        print e


def get_last(symbol):
    wb = load_workbook('Book1.xlsm',True,True,True,False,True)
    ws = wb['Sheet1']
    i = excel_rows[symbol]
    if i >=4:
        cell = ws['B{0}'.format(i)]
        return cell.value
    else:
        return None
    wb.close()


def get_bid(symbol):
    wb = load_workbook('Book1.xlsm', True, True, True, False, True)
    ws = wb['Sheet1']
    i = excel_rows[symbol]
    if i >= 4:
        cell = ws['F{0}'.format(i)]
        return cell.value
    else:
        return None
    wb.close()


def get_ask(symbol):
    wb = load_workbook('Book1.xlsm', True, True, True, False, True)
    ws = wb['Sheet1']
    i = excel_rows[symbol]
    if i >= 4:
        cell = ws['G{0}'.format(i)]
        return cell.value
    else:
        return None
    wb.close()

# build_dictionary()
# print get_last('GLD')